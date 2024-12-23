import ssl
import socket
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from cryptography import x509
from cryptography.hazmat.backends import default_backend
import openpyxl 
from openpyxl.styles import Alignment, Border, Side, PatternFill
import os
import argparse
import re

timeout = 10

# Function to get SSL certificate information for a hostname
def get_ssl_info(hostinfo : tuple):
    hostname = hostinfo[0]
    port = int(hostinfo[1]) if ( len( hostinfo ) > 1 ) else 443

    context = ssl.create_default_context()
    conn = context.wrap_socket(socket.socket(socket.AF_INET), server_hostname=hostname)

    try:
        conn.settimeout(timeout)
        conn.connect((hostname, port))
        cert = conn.getpeercert()
        cert_bin = conn.getpeercert(binary_form = True) #certificate in binary format
        cert2 = x509.load_der_x509_certificate(cert_bin, default_backend())

        # Expiration Date
        not_after = datetime.strptime(cert.get("notAfter"), '%b %d %H:%M:%S %Y GMT')

        # Date 
        today = datetime.today()

        # Days left
        remaining_days = (not_after - today).days

        #public key size
        public_key = cert2.public_key()
        key_size = public_key.key_size if hasattr(public_key, "key_size") else "Unknown"

        return {
            "Hname" : hostname,
            "Rdays" : remaining_days,
            "Ksize" : key_size,
            "Date" : not_after
        }
    except Exception as e:
        return f"Hostname: {hostname}\nError: {str(e)}\n"
    finally:
        conn.close()

# Read hostanmes
def read_hostnames(file_path):
    with open(file_path, 'r') as file:
        return [tuple( line.strip().split() ) for line in file]

def process_certificates(hostnames, max_workers=20):
    results = []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = executor.map(get_ssl_info, hostnames)
    return results

def save_to_excel(result, args):
    try: 
        hit = 0
        miss = 0
        list_misses = []

        #load excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.append(["Error?", "Hostname", "Remaining Days", "Public Key Size", "Expiration Date"])

        #row where insert data
        start_row = sheet.max_row + 1
        
        for i, entry in enumerate(result, start = start_row):
            if "Hname" in entry:      #data exists
                sheet.cell(row = i, column = 2, value = entry["Hname"])
                sheet.cell(row = i, column = 3, value = entry["Rdays"])
                sheet.cell(row = i, column = 4, value = entry["Ksize"])
                sheet.cell(row = i, column = 5, value = entry["Date"])
                hit = hit + 1
                rdays = entry["Rdays"]
                cellD = sheet.cell(row = i, column = 3)
                cellP = sheet.cell(row = i, column = 2)
                cmpNformat(rdays, cellD, cellP)
            else:                       #no data
                match = re.search(r"Hostname:\s*(\S+)", entry)
                if match:
                    hostname = match.group(1)
                    sheet.cell(row = i, column = 2, value = hostname)
                    sheet.cell(row = i, column = 1, value = "ERROR")
                miss = miss + 1
                list_misses.append(entry)

        sheet.cell(row = 4, column = 8, value = "HITS")
        sheet.cell(row = 4, column = 9, value = "MISS")
        sheet.cell(row = 5, column = 8, value = str(hit))
        sheet.cell(row = 5, column = 9, value = str(miss))

        total = hit+miss
        format_cells(sheet, total)

        #save changes
        name = generate_excel_name()
        workbook.save(name)

        print ("Data saved correctly into excel file with ", hit, " hits and ", miss, " misses")
        
        if args.errors:
            print("Misses are:\n")
            for i in range (len(list_misses)):
                print("[", i, "]:\n")
                print(list_misses[i])
                print("\n")
        else: 
            error_log = "misses.txt"
            f = open(error_log, 'w')
            f.write("Misses are:\n")
            for i in range (len(list_misses)):
                f.write("[")
                f.write(str(i))
                f.write("]")
                f.write(list_misses[i])
                f.write("\n")

    except Exception as e:
        print(f"Error saving Excel file: {e}")

def format_cells(sheet, total):
    h = "center"
    v = "center"
    black = "00000000"
    grey = "00D9D9D9"
    m = Side(border_style="thin", color = black)

    #adjust dimensions of the cells
    sheet.column_dimensions['B'].width = 27     #hostname column
    sheet.column_dimensions['C'].width = 15     #remaining days column
    sheet.column_dimensions['D'].width = 15     #public key size column
    sheet.column_dimensions['E'].width = 20     #expiration date column

    #MANUAL ALIGNMENT & COLOR
    sheet["H4"].alignment = Alignment(horizontal = h, vertical = v)
    sheet["I4"].alignment = Alignment(horizontal = h, vertical = v)
    sheet["H4"].border = Border(top = m, left = m, right = m, bottom = m)
    sheet["I4"].border = Border(top = m, left = m, right = m, bottom = m)
    sheet["H4"].fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")
    sheet["I4"].fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")
    sheet["H5"].alignment = Alignment(horizontal = h, vertical = v)
    sheet["I5"].alignment = Alignment(horizontal = h, vertical = v)
    sheet["H5"].border = Border(top = m, left = m, right = m, bottom = m)
    sheet["I5"].border = Border(top = m, left = m, right = m, bottom = m)

    sheet["A1"].fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")
    sheet["B1"].fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")
    sheet["C1"].fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")
    sheet["D1"].fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")
    sheet["E1"].fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")


    #AUTO ALIGNMENT & COLOR
    for r in range(1, total+2):
        for c in range(1, 6):
            cell = sheet.cell(row = r, column = c)
            cell.alignment = Alignment(horizontal= h, vertical= v)
            cell.border = Border(top = m, left = m, right = m, bottom = m)

#compare and format
def cmpNformat(rdays, cellD, cellP):
    fine = 100 
    attention = 50
    green = "0088C080"
    yellow = "00F7D570"
    red = "00F28C8C"
    if rdays >= fine:
        cellD.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
        cellP.fill = PatternFill(start_color=green, end_color=green, fill_type="solid")
    else:
        if rdays >= attention and rdays < fine:
            cellD.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
            cellP.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")
        else: 
            cellD.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")
            cellP.fill = PatternFill(start_color=red, end_color=red, fill_type="solid")

def generate_excel_name(base_name="certificates"):
    current_date = datetime.now().strftime("%Y-%m-%d")
    return f"{base_name}_{current_date}.xlsx"